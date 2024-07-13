import openpyxl as xl
from openpyxl.chart import Reference,BarChart
wb = xl.load_workbook('sheet.xlsx')
sheet = wb['Sheet1']
cell = sheet.cell(1,1)
for row in range (2,sheet.max_row+1):
    cell =  sheet.cell(row,3)
    correct = cell.value * 0.9
    corrected_cell = sheet.cell(row,4)
    corrected_cell.value = correct
values = Reference(sheet,
          min_row=2,
          max_row=sheet.max_row
          ,min_col=4,
          max_col=4
          )
chart =  BarChart()
chart.add_data(values)
sheet.add_chart(chart,'e2')
wb.save('sheet2.xlsx')
